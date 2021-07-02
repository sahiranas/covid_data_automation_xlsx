import openpyxl as xl
from openpyxl.chart import BarChart, Series, Reference


def covid_data_automation(filename):
    wb = xl.load_workbook(filename)

    def tpr_calculator(each_row, positive, test, tpr):
        tpr_cell = sheet.cell(each_row, tpr)
        no_of_test_cell = sheet.cell(each_row, test)
        positive_cell = sheet.cell(each_row, positive)
        tpr_cell.value = (positive_cell.value / no_of_test_cell.value) * 100

    def cumulative(each_row, cell1, cell2, result):
        cell1 = sheet.cell(each_row, cell1)
        cell2 = sheet.cell(each_row, cell2)
        result_cell = sheet.cell(each_row, result)
        result_cell.value = cell1.value + cell2.value

    def negative(each_row, test, positive, negative_reports):
        test = sheet.cell(each_row, test)
        positive = sheet.cell(each_row, positive)
        result_cell = sheet.cell(each_row, negative_reports)
        result_cell.value = test.value - positive.value

    sheet_names = wb.sheetnames
    no_of_sheets = len(sheet_names)

    for each in sheet_names:
        sheet = wb[each]
        for row in range(4, sheet.max_row + 1):
            negative(row, test=4, positive=3, negative_reports=5)
            negative(row, test=8, positive=7, negative_reports=9)
            tpr_calculator(row, 3, 4, 6)
            tpr_calculator(row, 7, 8, 10)
            cumulative(row, cell1=3, cell2=7, result=11)
            cumulative(row, cell1=4, cell2=8, result=12)
            cumulative(row, cell1=5, cell2=9, result=13)

    sheets = wb["Sheet1"]
    chart1 = BarChart()

    chart1.type = "col"
    chart1.style = 10

    chart1.title = "Positive Cases"
    chart1.y_axis.title = 'Numbers'
    chart1.x_axis.title = 'Local body'

    data = Reference(sheets, min_col=3, max_col=6, min_row=3, max_row=7)
    cats = Reference(sheets, min_col=1, min_row=4, max_row=7)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheets.add_chart(chart1, "P4")

    wb.save(filename)
